import openpyxl
import numpy
import pickle

wb = openpyxl.load_workbook("./teachdata.xlsx")
ws = wb["Sheet1"]

header_cells = ws[1]

clothes_list = []
for row in ws.iter_rows(min_row=2):
    row_dic = {}
    for k,v in zip(header_cells,row):
        if v.value is None:
            break
        if k.value == 'image_id':
            row_dic[k.value] = "./dataset/" + str(v.value) + ".jpg"
        else:
            row_dic[k.value] = v.value
    clothes_list.append(row_dic)

clothes_list.pop(-1)

print(clothes_list)

f = open('clothes.pickle',mode='wb')
pickle.dump(clothes_list,f)
f.close